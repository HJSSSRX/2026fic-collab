"""Generate Excel report for server_analyst progress (2026FIC team competition).

Per user rule: every report/handoff MUST include .xlsx with columns:
题号, 题目, 答案, 解析, 证据位置, 独立解出标记
"""
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = Path(r"E:\ffffff-JIANCAI\2026FIC团体赛\case\server\REPORT_server.xlsx")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "服务器取证答题"

# Header
headers = ["题号", "题目", "状态", "答案", "解析", "证据位置", "独立解出"]
ws.append(headers)

# Style header row
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="2E5C8A")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
thin = Side(border_style="thin", color="888888")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col_idx, _ in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

# Data rows: (题号, 题目, 状态, 答案, 解析, 证据位置, 独立解出)
rows = [
    # ---- 服务器组 (Q1-Q17) ----
    ("Q1", "操作系统版本",
     "✓ 已解",
     "Debian 13 (trixie), VERSION_ID=13",
     "从/etc/os-release提取(LV root部分明文区域); PRETTY_NAME=\"Debian GNU/Linux 13 (trixie)\", DEBIAN_VERSION_FULL=13.0",
     "F-S002; 检材3-2 LV root /etc/os-release扇区",
     "是"),
    
    ("Q2", "操作系统安装时间",
     "⏳ 阻塞",
     "(待LV root解密)",
     "需要解密LV root后从/var/log/installer/或/etc/machine-id文件创建时间获取. LV root疑似dm-crypt加密(PE数据全零).",
     "F-S005 blocker",
     "否"),
    
    ("Q3", "ngrok访问域名",
     "✓ 已解",
     "blemish-junior-unengaged.ngrok-free.dev",
     "从ngrok agent进程内存/配置文件残留提取; 该域名是ngrok-free.dev的随机三段式子域. ngrok将本地3000端口暴露至此域名.",
     "F-S006; 检材3-1 ngrok config blob (offset 0x709A9EE68等多处)",
     "是"),
    
    ("Q4", "ngrok具体版本号",
     "⏳ 阻塞",
     "(待解)",
     "已确认ngrok agent v3存在(/usr/local/bin/ngrok). 完整版本号需读取binary或config; LV root加密阻塞.",
     "F-S004 evidence",
     "否"),
    
    ("Q5", "网站后台管理入口文件名",
     "✓ 已解",
     "admin1.php",
     "MacCMS默认入口admin.php被改名为admin1.php作为安全混淆. 从nginx access.log多条URL确认: /admin1.php/admin/index/index.html, /admin1.php/admin/system/config.html等. 服务器内网IP=192.168.226.128",
     "F-S008; 检材3-1 nginx access logs (~30GB区域)",
     "是"),
    
    ("Q6", "网站名称",
     "✓ 已解",
     "免费短视频分享大全 - 大中国",
     "从MacCMS缓存PHP配置文件解析: 'site_name' => '免费短视频分享大全 - 大中国'",
     "F-S016; 检材3-2 offset 0x267D701CE (maccms.site cache)",
     "是"),
    
    ("Q7", "网站主域名",
     "✓ 已解",
     "www.2026fic.forensix",
     "从MacCMS缓存配置: 'site_url' => 'www.2026fic.forensix' (FIC比赛专用伪TLD .forensix). WAP域名: wap.2026fic.forensix",
     "F-S016; 检材3-2 offset 0x267D701CE",
     "是"),
    
    ("Q8", "网站ICP备案号",
     "✓ 已解",
     "icp1919810",
     "从MacCMS缓存配置: 'site_icp' => 'icp1919810'. 注意: 1919810是网络meme数字, 显然伪造.",
     "F-S016; 检材3-2 offset 0x267D701CE",
     "是"),
    
    ("Q9", "其它待解题",
     "⏳ 待评估",
     "(取决于具体题目)",
     "Q9题目内容需从role_prompt确认",
     "n/a",
     "—"),
    
    ("Q10", "MacCMS伪静态规则配置文件SM3哈希",
     "🔍 候选",
     "(候选 maccms.conf nginx版)",
     "找到MacCMS伪静态文件目录, 含 maccms.conf(nginx) / .htaccess(apache) / web.config(IIS) / httpd.ini(ISAPI). nginx版规则块: location/{if(!-e \\$request_filename){rewrite ^/index.php... rewrite ^/admin.php... 实际部署时admin.php应改为admin1.php才一致(否则后台无法访问), 故部署版的SM3≠源码版. 完整答案需LV root解密后读取/etc/nginx/conf.d/maccms.conf的实际部署文件",
     "F-S014; 检材3-1 offset 0x6FF98B8DF",
     "部分"),
    
    ("Q11", "数据库IP",
     "✓ 已解",
     "10.0.3.100",
     "从etcd/topology元数据(TiDB集群): /topology/tidb/10.0.3.100:4000/info",
     "F-S010; 检材3-1 多处offset (TiKV元数据区)",
     "是"),
    
    ("Q12", "数据库容器/部署技术",
     "✓ 已解",
     "TiUP (TiDB集群部署工具)",
     "deploy_path:/root/.tiup/components/tidb/v7.5.0; TiUP集群名=aamac. 注: TiUP是PingCAP官方TiDB集群管理工具(非Docker), 通过systemd管理TiDB+TiKV+PD+Grafana",
     "F-S011; 检材3-1 TiKV元数据 + tiup配置目录",
     "是"),
    
    ("Q13", "4000端口备份数据库版本",
     "✓ 已解",
     "TiDB v7.5.0",
     "从topology元数据JSON: {version:v7.5.0, git_hash:069631e2ecfedc000ffb92c67207bea81380f020, ip:10.0.3.100, status_port:10080}",
     "F-S010; 检材3-1 多处offset (TiKV元数据区)",
     "是"),
    
    ("Q14", "其它待解题",
     "⏳ 待评估",
     "(取决于具体题目)",
     "Q14题目内容需确认",
     "n/a",
     "—"),
    
    ("Q15", "其它待解题",
     "⏳ 待评估",
     "(取决于具体题目)",
     "Q15题目内容需确认",
     "n/a",
     "—"),
    
    ("Q16", "未被使用的文件系统",
     "✓ 推断",
     "A. NTFS",
     "已确认在用: BTRFS(LV data UUID 2fe53132-155e-c444-b224-e29cb4201c0e), XFS(Docker overlay), LVM(系统盘和数据盘), swap(检材3-2 sda1). EXT4内核模块/工具引用存在但无明确fstab挂载点. NTFS无任何痕迹(纯Linux服务器). 完整确认需解密LV root后查/etc/fstab",
     "F-S012; 综合多处证据",
     "是"),
    
    ("Q17", "已安装的数据库服务列表",
     "✓ 已解",
     "TiDB (含TiKV + TiFlash + PD)",
     "完整集群:`TiDB at 10.0.3.100:4000 (status_port=10080); TiKV at 10.0.3.100:20160; TiFlash at 10.0.3.100:20170; PD at 10.0.3.100:3930; ng-monitoring at 0.0.0.0:12020. Grafana仪表盘存在(tikv_fast_tune.json等). 选项含MySQL/PostgreSQL/MariaDB/TiDB则选TiDB",
     "F-S013; 检材3-1 TiKV元数据多处",
     "是"),
    
    # ---- 互联网组 (跨题) ----
    ("互联网Q1", "售卖卡密的Telegram群组",
     "✓ 已解",
     "FIC_2026 (https://t.me/FIC_2026)",
     "从MacCMS支付配置: pay.card.url='https://t.me/FIC_2026'. 这是网站后台['配置/在线支付配置/卡密设置/销售网址']的实际值. 区别于maccms_channel(MacCMS官方频道)",
     "F-S015; 检材3-2 offset 0x267D701CE",
     "是"),
    
    ("互联网Q2", "备份数据库视频图片文件名",
     "🔍 候选",
     "(候选: 见解析)",
     "已知: 1) site_banner配置3张图(upload/site/20260415-1/): 194647b41e567e1dc60458fcfb37f337.jpg / 3300bfeea11e342f23576fb25d281e0e.jpg / b0501209bbfb4c35a0c8edb37a7b3a11.jpg. 2) nginx access日志中16个 vod 图片(upload/vod/20260415-1/HASH.jpg)全部返回404(磁盘上不存在). 3) backup_path='./application/data/backup/database/' 但磁盘上无actual SQL INSERT INTO mac_vod记录(数据已迁到TiDB,二进制存储). 完整答案需解密LV data + TiDB备份提取",
     "F-S017; 检材3-2 offset 0x267D701CE + nginx logs",
     "部分"),
]

for r in rows:
    ws.append(r)

# Style data rows
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.border = border
        cell.alignment = left_wrap
        cell.font = Font(size=10)

# Color status column
status_fills = {
    "✓ 已解": PatternFill("solid", fgColor="C6EFCE"),
    "✓ 推断": PatternFill("solid", fgColor="DDEBF7"),
    "🔍 候选": PatternFill("solid", fgColor="FFEB9C"),
    "⏳ 阻塞": PatternFill("solid", fgColor="FFC7CE"),
    "⏳ 待评估": PatternFill("solid", fgColor="EDEDED"),
}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
    for cell in row:
        if cell.value in status_fills:
            cell.fill = status_fills[cell.value]
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Column widths
widths = {"A": 12, "B": 28, "C": 12, "D": 35, "E": 60, "F": 35, "G": 10}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# Row heights
for r in range(2, ws.max_row + 1):
    ws.row_dimensions[r].height = 75

# Freeze top row
ws.freeze_panes = "A2"

# Add summary sheet
ws2 = wb.create_sheet("汇总")
ws2.append(["类别", "数量"])
ws2.append(["✓ 已解 (含推断)", sum(1 for r in rows if r[2].startswith("✓"))])
ws2.append(["🔍 候选/部分", sum(1 for r in rows if r[2].startswith("🔍"))])
ws2.append(["⏳ 阻塞/待评估", sum(1 for r in rows if r[2].startswith("⏳"))])
ws2.append(["总题数(已知)", len(rows)])
ws2.append([])
ws2.append(["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
ws2.append(["分析角色", "server_analyst"])
ws2.append(["证据基础", "检材3-1.E0X (sdb, 60GB) + 检材3-2.E0X (sda, 60GB)"])
ws2.append(["关键阻塞", "LV root疑似dm-crypt加密(PE数据全零), 阻塞: Q2安装时间, Q4 ngrok版本, Q10 SM3完整, 互联网Q2精确文件名"])
ws2.column_dimensions["A"].width = 25
ws2.column_dimensions["B"].width = 90
for row in ws2.iter_rows():
    for cell in row:
        cell.alignment = left_wrap
        cell.border = border

wb.save(OUT)
print(f"Report saved: {OUT}")
print(f"Rows: {ws.max_row - 1} questions analyzed")

# Print summary to stdout
print("\n=== Summary ===")
solved = [r for r in rows if r[2].startswith("✓")]
candidate = [r for r in rows if r[2].startswith("🔍")]
blocked = [r for r in rows if r[2].startswith("⏳")]
print(f"  ✓ Solved/Inferred: {len(solved)}")
for r in solved:
    print(f"    {r[0]}: {r[3][:60]}")
print(f"  🔍 Candidate: {len(candidate)}")
for r in candidate:
    print(f"    {r[0]}")
print(f"  ⏳ Blocked: {len(blocked)}")
for r in blocked:
    print(f"    {r[0]}")
